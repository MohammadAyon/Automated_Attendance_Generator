"""
attendance_core.py
-------------------
Builds the multi-sheet attendance workbook (Settings, Normalized Data,
Summary, one tab per employee, Instructions) directly from the raw
punch-log CSV export -- matching the layout of Attendance_Automated_Aug2026.xlsx.

No template .xlsx is needed anymore; every sheet is styled from scratch.

Month-specific settings (standard hours, weekly off day, public holidays)
live in attendance_settings.json, created automatically next to this script
(or next to the .exe, when frozen) on first run. Edit that JSON file each
month -- no code changes or rebuild needed, which matters once this is
packaged as a standalone executable other people run.
"""

import calendar
import json
import os
import re
import sys
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Settings file (holidays, standard hours, weekly off day)
# ---------------------------------------------------------------------------
DEFAULT_TARGET_HOURS = 9
DEFAULT_WEEKLY_OFF = "Friday"
DEFAULT_HOLIDAYS = {
    "2026-08-05": "Holiday",
    "2026-08-26": "Holiday",
}


def app_dir():
    """Folder the .exe (or this .py, when not frozen) lives in -- so the
    settings file sits next to whichever one the person actually has."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


SETTINGS_FILE = os.path.join(app_dir(), "attendance_settings.json")


def _default_settings_dict():
    return {
        "target_hours": DEFAULT_TARGET_HOURS,
        "weekly_off": DEFAULT_WEEKLY_OFF,
        "holidays": dict(DEFAULT_HOLIDAYS),
    }


def load_settings(log=lambda m: None):
    """Reads attendance_settings.json next to the exe/script; creates it
    with defaults on first run. Returns (target_hours, weekly_off,
    holidays) where holidays maps datetime.date -> reason string. Never
    raises -- falls back to built-in defaults and logs a warning instead."""
    if not os.path.exists(SETTINGS_FILE):
        raw = _default_settings_dict()
        try:
            with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(raw, f, indent=2)
            log(f"Created {SETTINGS_FILE} with default settings. "
                f"Edit the \"holidays\" section there each month -- no rebuild needed.")
        except OSError as e:
            log(f"WARNING: couldn't create settings file ({e}); using built-in defaults.")
    else:
        try:
            with open(SETTINGS_FILE, encoding="utf-8") as f:
                raw = json.load(f)
        except (OSError, json.JSONDecodeError) as e:
            log(f"WARNING: couldn't read {SETTINGS_FILE} ({e}); using built-in defaults.")
            raw = _default_settings_dict()

    target_hours = raw.get("target_hours", DEFAULT_TARGET_HOURS)
    weekly_off = raw.get("weekly_off", DEFAULT_WEEKLY_OFF)
    holidays = {}
    for date_str, reason in raw.get("holidays", {}).items():
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d").date()
            holidays[d] = reason
        except ValueError:
            log(f"WARNING: skipping unreadable holiday date '{date_str}' "
                f"in {SETTINGS_FILE} (use YYYY-MM-DD).")
    return target_hours, weekly_off, holidays

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
BLACK = PatternFill("solid", fgColor="FF000000")
GRAY_DATA = PatternFill("solid", fgColor="FFE7E6E6")
WHITE_BOLD = Font(bold=True, color="FFFFFFFF")
RED_BOLD = Font(bold=True, color="FFFF0000")
GRAY_NOTE = Font(color="FF808080")
CENTER = Alignment(horizontal="center", vertical="center")


def title_font(size):
    return Font(bold=True, color="FFFFFFFF", size=size)


def style_range(ws, cell_range, fill=None, font=None, alignment=None, number_format=None):
    for row in ws[cell_range]:
        for cell in row:
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment
            if number_format is not None:
                cell.number_format = number_format


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# ---------------------------------------------------------------------------
# CSV parsing
# ---------------------------------------------------------------------------
NAME_RE = re.compile(r"^'")
DATE_RE = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")
TIME_RE = re.compile(r"^\d{1,2}:\d{2}:\d{2}$")


def parse_csv(path):
    """Returns {employee: {date: [time, time, ...]}} preserving CSV order."""
    employees = {}
    cur_emp = None
    cur_date = None

    with open(path, encoding="utf-8-sig", errors="replace") as f:
        lines = [l.strip() for l in f if l.strip()]

    for line in lines[1:]:  # skip "Row Labels" pivot header
        if line == "Grand Total":
            break
        if NAME_RE.match(line):
            cur_emp = line[1:]
            employees.setdefault(cur_emp, {})
            cur_date = None
        elif DATE_RE.match(line):
            cur_date = datetime.strptime(line, "%m/%d/%Y").date()
            employees.setdefault(cur_emp, {})[cur_date] = []
        elif TIME_RE.match(line):
            h, m, s = (int(x) for x in line.split(":"))
            if cur_emp is not None and cur_date is not None:
                employees[cur_emp][cur_date].append(
                    datetime.strptime(line, "%H:%M:%S").time())

    return employees


def detect_month(employees):
    all_dates = [d for days in employees.values() for d in days.keys()]
    if not all_dates:
        raise ValueError("No dated punch records found in the CSV.")
    from collections import Counter
    counts = Counter((d.year, d.month) for d in all_dates)
    return counts.most_common(1)[0][0]


def safe_sheet_name(name, used):
    base = re.sub(r"[\[\]:\*\?/\\]", "", name).strip()[:31] or "Employee"
    candidate = base
    i = 2
    while candidate in used:
        suffix = f" ({i})"
        candidate = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate)
    return candidate


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def build_settings_sheet(wb, year, month, holidays, target_hours, weekly_off):
    ws = wb.create_sheet("Settings")
    ws.merge_cells("A1:D1")
    ws["A1"] = "Attendance Automation Settings"
    style_range(ws, "A1:D1", fill=BLACK, font=title_font(14))

    ws["A3"], ws["B3"] = "Setting", "Value"
    ws["D3"], ws["E3"] = "Holiday Date", "Reason"
    style_range(ws, "A3:B3", fill=BLACK, font=WHITE_BOLD)
    style_range(ws, "D3:E3", fill=BLACK, font=WHITE_BOLD)

    ws["A4"] = "Report Month"
    ws["B4"] = datetime(year, month, 1)
    ws["B4"].number_format = "mmmm yyyy"
    ws["A5"] = "Standard Hours / Attended Day"
    ws["B5"] = target_hours
    ws["A6"] = "Weekly Off Day"
    ws["B6"] = weekly_off
    ws["A7"] = "Holiday Dates"
    ws["B7"] = ", ".join(d.isoformat() for d in sorted(holidays)) if holidays else "(none)"

    for i, d in enumerate(sorted(holidays), start=4):
        ws.cell(row=i, column=4, value=datetime(d.year, d.month, d.day)).number_format = "m/d/yyyy"
        ws.cell(row=i, column=5, value=holidays[d])

    set_widths(ws, {"A": 27, "B": 24, "D": 17, "E": 20})
    return ws


def build_normalized_sheet(wb, employees):
    ws = wb.create_sheet("Normalized Data")
    headers = ["Employee", "Date", "Punch Time", "Punch #", "First/Last Logic"]
    ws.append(headers)
    style_range(ws, "A1:E1", fill=BLACK, font=WHITE_BOLD)

    row = 2
    for name, by_date in employees.items():
        for d in sorted(by_date):
            times = sorted(by_date[d])
            logic_text = ("Single punch = incomplete" if len(times) == 1
                          else "First punch = Check-in; Last punch = Check-out")
            for i, t in enumerate(times, start=1):
                ws.cell(row=row, column=1, value=name)
                c = ws.cell(row=row, column=2, value=datetime(d.year, d.month, d.day))
                c.number_format = "m/d/yyyy"
                c2 = ws.cell(row=row, column=3, value=datetime.combine(d, t))
                c2.number_format = "h:mm:ss AM/PM"
                ws.cell(row=row, column=4, value=i)
                ws.cell(row=row, column=5, value=logic_text)
                row += 1

    set_widths(ws, {"A": 24, "B": 13, "C": 22, "D": 10, "E": 43})
    return ws


def build_employee_sheet(wb, sheet_name, display_name, punches_by_date,
                          year, month, days_in_month, holidays, weekly_off):
    ws = wb.create_sheet(sheet_name)

    ws.merge_cells("B5:I5")
    ws["B5"] = "Employee Attendance Summary"
    style_range(ws, "B5:I5", fill=BLACK, font=title_font(14),
                alignment=Alignment(horizontal="left", vertical="center"))

    ws.merge_cells("B8:C8")
    ws["B8"] = "Employee Name"
    ws["B8"].font = Font(bold=True)
    ws.merge_cells("D8:F8")
    ws["D8"] = display_name
    ws["D8"].font = Font(bold=True)
    ws["G8"] = "DATE:"
    ws["G8"].font = Font(bold=True)
    ws.merge_cells("H8:I8")
    ws["H8"] = datetime(year, month, days_in_month)
    ws["H8"].font = Font(bold=True)
    ws["H8"].alignment = Alignment(horizontal="center")
    ws["H8"].number_format = "d-mmm-yy"

    headers = ["DATE", "STATUS", "CHECK-IN", "CHECK-OUT", "WORKING HOURS", "Remarks"]
    for col, text in zip("BCDEFG", headers):
        ws[f"{col}11"] = text
    style_range(ws, "B11:G11", fill=BLACK, font=WHITE_BOLD, alignment=CENTER)
    ws.merge_cells("K11:M11")
    ws["K11"] = "DAILY"
    style_range(ws, "K11:M11", fill=BLACK, font=WHITE_BOLD,
                alignment=Alignment(horizontal="center"))

    first_row = 12
    for day in range(1, 32):
        row = first_row + day - 1
        if day > days_in_month:
            continue

        d = date(year, month, day)
        is_weekly_off = d.strftime("%A") == weekly_off
        is_holiday = d in holidays
        punches = sorted(punches_by_date.get(d, []))

        style_range(ws, f"B{row}:G{row}", fill=GRAY_DATA, alignment=CENTER)

        if is_weekly_off or is_holiday:
            ws[f"B{row}"] = "Holiday" if is_holiday else weekly_off
            style_range(ws, f"B{row}:G{row}", fill=BLACK, font=WHITE_BOLD, alignment=CENTER)
            ws[f"B{row}"].font = RED_BOLD
        else:
            ws[f"B{row}"] = datetime(year, month, day)
            ws[f"B{row}"].number_format = "m/d/yy"
            if punches:
                ws[f"C{row}"] = "Attended"
                checkin, checkout = punches[0], punches[-1]
                ws[f"D{row}"] = datetime.combine(d, checkin)
                ws[f"D{row}"].number_format = "h:mm AM/PM"
                if len(punches) >= 2:
                    ws[f"E{row}"] = datetime.combine(d, checkout)
                    ws[f"E{row}"].number_format = "h:mm AM/PM"
                else:
                    ws[f"G{row}"] = "Incomplete: single punch"
            else:
                ws[f"C{row}"] = "On Leave"
                ws[f"G{row}"] = "No punch record"

        ws[f"F{row}"] = (f'=IF(C{row}<>"Attended","",'
                          f'IF(OR(D{row}="",E{row}=""),0,E{row}-D{row}))')
        ws[f"F{row}"].number_format = "[h]:mm"

    last_row = first_row + min(days_in_month, 31) - 1

    ws["K12"], ws["M12"] = "Total Office Days", f"=COUNT(B{first_row}:B{last_row})"
    ws["K13"], ws["M13"] = "Total Attendance", f'=COUNTIF(C{first_row}:C{last_row},"Attended")'
    ws["K14"], ws["M14"] = "Total Leave", f'=COUNTIF(C{first_row}:C{last_row},"On Leave")'
    style_range(ws, "K12:K14", fill=BLACK, font=WHITE_BOLD)

    ws.merge_cells("K16:M16")
    ws["K16"] = "WORKING HOURS"
    style_range(ws, "K16:M16", fill=BLACK, font=WHITE_BOLD,
                alignment=Alignment(horizontal="center"))

    ws["K17"] = "Total Hours\n[Attendance *\n9H/PD]"
    ws["K17"].fill, ws["K17"].font = BLACK, WHITE_BOLD
    ws["M17"] = "=M13*Settings!B5"
    ws["M17"].number_format = "0.0"

    ws["K19"] = "Working Hours"
    ws["K19"].fill, ws["K19"].font = BLACK, WHITE_BOLD
    ws["M19"] = f"=SUM(F{first_row}:F{last_row})"
    ws["M19"].number_format = "[h]:mm"

    ws.merge_cells("K22:N22")
    ws["K22"] = '*Note: All absent days have been counted as "On Leave."'
    ws["K22"].font = GRAY_NOTE

    set_widths(ws, {"A": 3, "B": 14, "C": 14, "D": 13, "E": 13, "F": 15,
                     "G": 30, "H": 7, "I": 7, "J": 3, "K": 21, "L": 3, "M": 16, "N": 3})


def build_summary_sheet(wb, sheet_names_by_employee):
    ws = wb.create_sheet("Summary")
    ws.merge_cells("A1:G1")
    ws["A1"] = "Employee Attendance Summary"
    style_range(ws, "A1:G1", fill=BLACK, font=title_font(15))

    headers = ["Employee", "Office Days", "Attendance", "Leave",
               "Target Hours", "Actual Working Hours", "Punch Records"]
    ws.append([None] * 7)  # keep row2 blank like the source layout
    ws.append(headers)
    style_range(ws, "A3:G3", fill=BLACK, font=WHITE_BOLD)

    row = 4
    for display_name, sheet_name in sheet_names_by_employee:
        q = sheet_name.replace("'", "''")
        ws.cell(row=row, column=1, value=display_name)
        ws.cell(row=row, column=2, value=f"='{q}'!M12")
        ws.cell(row=row, column=3, value=f"='{q}'!M13")
        ws.cell(row=row, column=4, value=f"='{q}'!M14")
        ws.cell(row=row, column=5, value=f"='{q}'!M17")
        ws.cell(row=row, column=6, value=f"='{q}'!M19")
        ws.cell(row=row, column=7,
                value=f"=COUNTIF('Normalized Data'!$A:$A,A{row})")
        row += 1

    set_widths(ws, {"A": 25, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 14})
    return ws


def build_instructions_sheet(wb, holidays, target_hours, weekly_off):
    ws = wb.create_sheet("Instructions")
    ws.merge_cells("A1:H1")
    ws["A1"] = "How the automation works"
    style_range(ws, "A1:H1", fill=BLACK, font=title_font(14))

    def format_day(d):
        if not hasattr(d, "strftime"):
            return str(d)
        return d.strftime("%d-%b-%Y").lstrip("0")

    if holidays:
        holiday_txt = " / ".join(
            format_day(d) for d in sorted(holidays)
        )
        holiday_line = f"5. {weekly_off}s are weekly off days and {holiday_txt} are configured holidays for this report."
    else:
        holiday_line = f"5. {weekly_off}s are weekly off days. No extra holidays are configured for this report."

    text = (
        "1. Import the one-column recordList.csv into the automation.\n"
        "2. The parser treats each employee line, date line, and time line as a hierarchy.\n"
        "3. For each employee/date: first punch becomes Check-in; last punch becomes Check-out.\n"
        "4. One-punch days are marked Attended but flagged as Incomplete.\n"
        f"{holiday_line}\n"
        "6. Days with no punches are marked On Leave, matching the example workbook's note.\n"
        "7. Each employee receives a separate report sheet plus the Summary and Normalized Data sheets.\n"
        f"8. Standard target is {target_hours} hours per attended day; edit attendance_settings.json to change it."
    )
    ws.merge_cells("A3:H12")
    ws["A3"] = text
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")

    set_widths(ws, {"A": 24, **{c: 15 for c in "BCDEFGH"}})
    return ws


# ---------------------------------------------------------------------------
# Top-level build
# ---------------------------------------------------------------------------
def build_workbook(csv_path, output_path, log=print):
    target_hours, weekly_off, holidays = load_settings(log)

    employees = parse_csv(csv_path)
    log(f"Found {len(employees)} employee(s) in the file.")

    year, month = detect_month(employees)
    days_in_month = calendar.monthrange(year, month)[1]
    log(f"Report period detected: {calendar.month_name[month]} {year}")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    build_settings_sheet(wb, year, month, holidays, target_hours, weekly_off)
    build_normalized_sheet(wb, employees)

    used_names = set()
    sheet_names_by_employee = []
    stats = {"employees": 0, "zero_punch_employees": [], "single_punch_days": 0,
             "early_checkins": [], "on_leave_days": 0}

    for name in sorted(employees.keys(), key=str.lower):
        sheet_name = safe_sheet_name(name, used_names)
        sheet_names_by_employee.append((name, sheet_name))
        by_date = employees[name]
        build_employee_sheet(wb, sheet_name, name, by_date, year, month,
                              days_in_month, holidays, weekly_off)
        stats["employees"] += 1
        if not by_date:
            stats["zero_punch_employees"].append(name)
        for d, times in by_date.items():
            if d.strftime("%A") == weekly_off or d in holidays:
                continue
            if len(times) == 1:
                stats["single_punch_days"] += 1
            if times and min(times).hour < 4:
                stats["early_checkins"].append((name, d.isoformat(), min(times).isoformat()))
        for day in range(1, days_in_month + 1):
            d = date(year, month, day)
            if d.strftime("%A") == weekly_off or d in holidays:
                continue
            if d not in by_date:
                stats["on_leave_days"] += 1

    build_summary_sheet(wb, sheet_names_by_employee)
    build_instructions_sheet(wb, holidays, target_hours, weekly_off)

    # Sheet order: Settings, Normalized Data, Summary, employees..., Instructions
    order = ["Settings", "Normalized Data", "Summary"] + \
            [s for _, s in sheet_names_by_employee] + ["Instructions"]
    wb._sheets = [wb[name] for name in order]

    wb.save(output_path)
    stats["month"] = f"{calendar.month_name[month]} {year}"
    log(f"Saved: {output_path}")
    return stats
